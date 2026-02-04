#!/usr/bin/env python3
"""
Script to compile all final_report-{model}.json files into a single Excel workbook.

Usage:
    python3 scripts/compile_final_reports.py \
        --reports_dir data/evaluation \
        --output_file data/evaluation/final_reports_comparison.xlsx
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List

try:
    import pandas as pd
except ImportError:
    print("Error: pandas library not found. Install with: pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl library not found. Install with: pip install openpyxl")
    sys.exit(1)


def find_final_reports(reports_dir: Path) -> Dict[str, Dict[str, Path]]:
    """
    Find all final_report-{model}.json files organized by repo and model.

    Args:
        reports_dir: Directory to search

    Returns:
        Dictionary mapping repo_name -> model_name -> file_path
    """
    reports = {}

    # Search for all final_report-*.json files
    for report_file in reports_dir.glob("**/final_report-*.json"):
        # Extract repo name (parent directory)
        repo_name = report_file.parent.name

        # Extract model name from filename
        filename = report_file.name
        if filename.startswith("final_report-") and filename.endswith(".json"):
            model_name = filename[len("final_report-"):-len(".json")]
        else:
            continue

        # Store the file path
        if repo_name not in reports:
            reports[repo_name] = {}
        reports[repo_name][model_name] = report_file

    return reports


def load_final_report(report_file: Path) -> Dict:
    """Load a final report JSON file."""
    try:
        with open(report_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Failed to load {report_file}: {e}", file=sys.stderr)
        return {}


def create_comparison_dataframe(reports: Dict[str, Dict[str, Path]]) -> pd.DataFrame:
    """
    Create a comparison dataframe from all reports.

    Args:
        reports: Dictionary mapping repo_name -> model_name -> file_path

    Returns:
        DataFrame with comparison data
    """
    rows = []

    for repo_name in sorted(reports.keys()):
        for model_name in sorted(reports[repo_name].keys()):
            report_file = reports[repo_name][model_name]
            report_data = load_final_report(report_file)

            if not report_data:
                continue

            # Calculate resolution rate
            total = report_data.get('total_instances', 0)
            resolved = report_data.get('resolved_instances', 0)
            resolution_rate = (resolved / total * 100) if total > 0 else 0

            row = {
                'Repository': repo_name,
                'Model': model_name,
                'Total Instances': total,
                'Submitted': report_data.get('submitted_instances', 0),
                'Completed': report_data.get('completed_instances', 0),
                'Resolved': resolved,
                'Unresolved': report_data.get('unresolved_instances', 0),
                'Resolution Rate (%)': resolution_rate,
                'Incomplete': report_data.get('incomplete_instances', 0),
                'Empty Patch': report_data.get('empty_patch_instances', 0),
                'Errors': report_data.get('error_instances', 0)
            }
            rows.append(row)

    return pd.DataFrame(rows)


def create_pivot_tables(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Create pivot tables for different metrics.

    Args:
        df: Main comparison dataframe

    Returns:
        Dictionary of pivot tables
    """
    pivot_tables = {}

    # Resolution Rate pivot (repo x model)
    if not df.empty:
        pivot_tables['Resolution Rate'] = df.pivot_table(
            index='Repository',
            columns='Model',
            values='Resolution Rate (%)',
            aggfunc='mean'
        ).round(2)

        # Resolved Instances pivot
        pivot_tables['Resolved Instances'] = df.pivot_table(
            index='Repository',
            columns='Model',
            values='Resolved',
            aggfunc='sum',
            fill_value=0
        ).astype(int)

        # Total Instances pivot
        pivot_tables['Total Instances'] = df.pivot_table(
            index='Repository',
            columns='Model',
            values='Total Instances',
            aggfunc='mean',
            fill_value=0
        ).astype(int)

    return pivot_tables


def create_summary_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a summary sheet with aggregate statistics per model.

    Args:
        df: Main comparison dataframe

    Returns:
        Summary dataframe
    """
    if df.empty:
        return pd.DataFrame()

    summary_rows = []

    for model in sorted(df['Model'].unique()):
        model_data = df[df['Model'] == model]

        total_instances = model_data['Total Instances'].sum()
        total_resolved = model_data['Resolved'].sum()
        resolution_rate = (total_resolved / total_instances * 100) if total_instances > 0 else 0

        summary_row = {
            'Model': model,
            'Total Repos': len(model_data),
            'Total Instances': total_instances,
            'Total Resolved': total_resolved,
            'Resolution Rate (%)': resolution_rate,
            'Avg Resolution Rate (%)': model_data['Resolution Rate (%)'].mean(),
            'Total Completed': model_data['Completed'].sum(),
            'Total Incomplete': model_data['Incomplete'].sum(),
            'Total Empty Patches': model_data['Empty Patch'].sum(),
            'Total Errors': model_data['Errors'].sum()
        }
        summary_rows.append(summary_row)

    summary_df = pd.DataFrame(summary_rows)
    summary_df = summary_df.round(2)

    return summary_df


def compile_reports(reports_dir: Path, output_file: Path):
    """
    Compile all final reports into an Excel workbook.

    Args:
        reports_dir: Directory containing final reports
        output_file: Output Excel file path
    """
    print(f"Searching for final reports in: {reports_dir}")
    reports = find_final_reports(reports_dir)

    if not reports:
        print("Error: No final report files found")
        sys.exit(1)

    total_reports = sum(len(models) for models in reports.values())
    print(f"Found {total_reports} final reports across {len(reports)} repositories")

    # Create main comparison dataframe
    print("Creating comparison dataframe...")
    df = create_comparison_dataframe(reports)

    if df.empty:
        print("Error: No valid data to compile")
        sys.exit(1)

    # Create pivot tables
    print("Creating pivot tables...")
    pivot_tables = create_pivot_tables(df)

    # Create summary
    print("Creating summary...")
    summary_df = create_summary_sheet(df)

    # Write to Excel
    print(f"\nCreating Excel workbook: {output_file}")
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write summary sheet
        if not summary_df.empty:
            print("Writing summary sheet...")
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Format summary sheet
            worksheet = writer.sheets['Summary']
            for idx, col in enumerate(summary_df.columns):
                max_length = max(
                    summary_df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = min(max_length + 2, 50)

        # Write pivot tables
        for sheet_name, pivot_df in pivot_tables.items():
            if not pivot_df.empty:
                print(f"Writing {sheet_name} sheet...")
                pivot_df.to_excel(writer, sheet_name=sheet_name)

                # Format sheet
                worksheet = writer.sheets[sheet_name]
                for idx in range(len(pivot_df.columns) + 1):  # +1 for index column
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = 20

        # Write detailed data
        print("Writing detailed data sheet...")
        df.to_excel(writer, sheet_name='Detailed Data', index=False)

        # Format detailed data sheet
        worksheet = writer.sheets['Detailed Data']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = min(max_length + 2, 50)

    print(f"\n✓ Successfully created Excel workbook")
    print(f"  - Summary sheet with per-model statistics")
    print(f"  - {len(pivot_tables)} pivot tables")
    print(f"  - Detailed data sheet")
    print(f"\nOutput file: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description="Compile all final_report-{model}.json files into a single Excel workbook"
    )
    parser.add_argument(
        "--reports_dir",
        type=Path,
        default=Path("data/evaluation"),
        help="Directory containing final report files (default: data/evaluation)"
    )
    parser.add_argument(
        "--output_file",
        type=Path,
        default=Path("data/evaluation/final_reports_comparison.xlsx"),
        help="Path to output Excel file (default: data/evaluation/final_reports_comparison.xlsx)"
    )

    args = parser.parse_args()

    # Validate input directory exists
    if not args.reports_dir.exists():
        print(f"Error: Directory not found: {args.reports_dir}", file=sys.stderr)
        sys.exit(1)

    # Compile reports
    compile_reports(args.reports_dir, args.output_file)


if __name__ == "__main__":
    main()
